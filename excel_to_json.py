from openpyxl import load_workbook
import json
wb=load_workbook("books.xlsx")
ws=wb.active
headers=[c.value for c in ws[1]]
rows=[]
for r in ws.iter_rows(min_row=2, values_only=True):
    rows.append(dict(zip(headers,r)))
with open("books.json","w",encoding="utf-8") as f:
    json.dump(rows,f,indent=2)
print("books.json updated!")
